#!/usr/bin/env python3
"""
YouTube Search Query Generator for Door Detection Data Collection
=================================================================
Reads the door_data_collection_plan.xlsx and generates structured
YouTube search queries with URL encoding, priority ordering,
language/quality variations, and export to JSON/CSV.

Usage:
    python youtube_query_generator.py                        # Generate all queries → JSON + CSV
    python youtube_query_generator.py --priority P1 P3 P4    # Only specific priorities
    python youtube_query_generator.py --min-yield 0.4        # Only high-yield queries
    python youtube_query_generator.py --expand               # Add language/quality/year variations
    python youtube_query_generator.py --max-results 20       # YouTube results per query
    python youtube_query_generator.py --format csv           # Output format (csv/json/both/urls)
    python youtube_query_generator.py --stats                # Print statistics only
"""

import argparse
import csv
import json
import os
import sys
from dataclasses import dataclass, field, asdict
from typing import Optional
from urllib.parse import quote_plus, urlencode

try:
    from openpyxl import load_workbook
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl --break-system-packages -q")
    from openpyxl import load_workbook


# ─────────────────────────────────────────────
# Data model
# ─────────────────────────────────────────────
@dataclass
class SearchQuery:
    term: str
    priority: str          # P1, P2, ..., P12
    priority_rank: int     # 1-12 for sorting
    category: str          # Sheet-level category name
    sub_category: str      # Sub-category within sheet
    est_videos: int
    yield_ratio: float
    est_images: float
    notes: str
    youtube_url: str = ""
    youtube_api_query: str = ""
    is_variation: bool = False
    variation_type: str = ""  # "original", "lang_ja", "quality_4k", etc.

    def __post_init__(self):
        if not self.youtube_url:
            self.youtube_url = self.build_youtube_url()
        if not self.youtube_api_query:
            self.youtube_api_query = self.term

    def build_youtube_url(self, max_results: int = 50) -> str:
        params = {
            "search_query": self.term,
            "sp": "CAASAhAB",  # Sort by relevance, filter: video only
        }
        return f"https://www.youtube.com/results?{urlencode(params)}"

    def build_api_params(self, max_results: int = 50) -> dict:
        """Parameters for YouTube Data API v3 search.list"""
        return {
            "part": "snippet",
            "q": self.term,
            "type": "video",
            "maxResults": min(max_results, 50),
            "order": "relevance",
            "videoDuration": "medium",  # 4-20 min — best for door content
        }


# ─────────────────────────────────────────────
# Sheet → Priority mapping
# ─────────────────────────────────────────────
SHEET_CONFIG = {
    "P1 Glass Doors Direct": {
        "priority": "P1",
        "rank": 1,
        "category": "Glass Doors — Direct Search",
    },
    "P2 Walking Tours": {
        "priority": "P2",
        "rank": 2,
        "category": "Walking Tours & Walkthroughs",
    },
    "P3 CCTV Security View": {
        "priority": "P3",
        "rank": 3,
        "category": "CCTV & Security Camera View",
    },
    "P4 Warehouse Industrial": {
        "priority": "P4",
        "rank": 4,
        "category": "Warehouse & Industrial",
    },
    "P5 POV Action": {
        "priority": "P5",
        "rank": 5,
        "category": "POV & Action",
    },
}

# P6-P12 are in a single sheet, identified by sub_category prefix
P6_P12_PRIORITY_MAP = {
    "P6 Real Estate": ("P6", 6, "Real Estate & Architecture"),
    "P7 Technical": ("P7", 7, "Technical & Installation"),
    "P8 Transit": ("P8", 8, "Transportation Hubs"),
    "P9 Retail": ("P9", 9, "Retail & Commercial"),
    "P10 Live Cam": ("P10", 10, "Live Cams & Streams"),
    "P11 Fails": ("P11", 11, "Fails & Compilations"),
    "P12 Residential": ("P12", 12, "Residential Doors"),
}


# ─────────────────────────────────────────────
# Excel reader
# ─────────────────────────────────────────────
def read_queries_from_excel(filepath: str) -> list[SearchQuery]:
    wb = load_workbook(filepath, data_only=True)
    queries = []

    # Process P1-P5 sheets (single priority per sheet)
    for sheet_name, config in SHEET_CONFIG.items():
        if sheet_name not in wb.sheetnames:
            print(f"  [WARN] Sheet '{sheet_name}' not found, skipping")
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=3, max_col=6, values_only=True):
            term, sub_cat, est_v, yld, notes, est_img = row
            if not term or not isinstance(term, str):
                continue
            queries.append(SearchQuery(
                term=term.strip(),
                priority=config["priority"],
                priority_rank=config["rank"],
                category=config["category"],
                sub_category=str(sub_cat or ""),
                est_videos=int(est_v or 0),
                yield_ratio=float(yld or 0),
                est_images=float(est_img or 0),
                notes=str(notes or ""),
                variation_type="original",
            ))

    # Process P6-P12 combined sheet
    combined_sheet = "P6-P12 Additional"
    if combined_sheet in wb.sheetnames:
        ws = wb[combined_sheet]
        for row in ws.iter_rows(min_row=3, max_col=6, values_only=True):
            term, sub_cat, est_v, yld, notes, est_img = row
            if not term or not isinstance(term, str):
                continue
            sub_cat_str = str(sub_cat or "")
            priority, rank, category = "P6", 6, "Other"
            for prefix, (p, r, c) in P6_P12_PRIORITY_MAP.items():
                if sub_cat_str.startswith(prefix):
                    priority, rank, category = p, r, c
                    break
            queries.append(SearchQuery(
                term=term.strip(),
                priority=priority,
                priority_rank=rank,
                category=category,
                sub_category=sub_cat_str,
                est_videos=int(est_v or 0),
                yield_ratio=float(yld or 0),
                est_images=float(est_img or 0),
                notes=str(notes or ""),
                variation_type="original",
            ))

    # Process Live Cam Sources sheet → convert to search queries
    live_sheet = "Live Cam Sources"
    if live_sheet in wb.sheetnames:
        ws = wb[live_sheet]
        for row in ws.iter_rows(min_row=3, max_col=5, values_only=True):
            source, typ, relevance, how_to, notes = row
            if not source or not isinstance(source, str):
                continue
            # Extract YouTube search term from "How to Use" column if possible
            search_term = _extract_live_cam_query(source, str(how_to or ""))
            if search_term:
                stars = str(relevance or "").count("★")
                queries.append(SearchQuery(
                    term=search_term,
                    priority="P10",
                    priority_rank=10,
                    category="Live Cams & Streams",
                    sub_category=str(typ or ""),
                    est_videos=200,
                    yield_ratio=max(0.05, stars * 0.05),
                    est_images=200 * max(0.05, stars * 0.05),
                    notes=f"[Live Source: {source}] {notes or ''}",
                    variation_type="live_cam",
                ))

    wb.close()
    return queries


def _extract_live_cam_query(source: str, how_to: str) -> Optional[str]:
    """Try to extract a usable YouTube search query from live cam source info."""
    # If 'how_to' contains a quoted search suggestion, use it
    if "'" in how_to:
        parts = how_to.split("'")
        for i in range(1, len(parts), 2):
            candidate = parts[i].strip()
            if len(candidate) > 5:
                return candidate
    # If source starts with "YouTube:", the rest is the query hint
    if source.lower().startswith("youtube:"):
        candidate = source.split(":", 1)[1].strip().strip("'\"")
        return candidate
    return None


# ─────────────────────────────────────────────
# Query expansion (language, quality, year)
# ─────────────────────────────────────────────
LANGUAGE_EXPANSIONS = {
    "ja": {  # Japanese
        "walking tour": "ウォーキングツアー",
        "shopping mall": "ショッピングモール",
        "automatic door": "自動ドア",
        "glass door": "ガラスドア",
        "entrance": "入口",
        "warehouse": "倉庫",
        "office tour": "オフィスツアー",
        "train station": "駅",
        "convenience store": "コンビニ",
    },
    "ko": {  # Korean
        "walking tour": "걷기 투어",
        "shopping mall": "쇼핑몰",
        "automatic door": "자동문",
        "glass door": "유리문",
        "entrance": "입구",
        "warehouse": "창고",
        "subway station": "지하철역",
    },
    "zh": {  # Chinese
        "walking tour": "步行游",
        "shopping mall": "购物中心",
        "automatic door": "自动门",
        "glass door": "玻璃门",
        "entrance": "入口",
        "warehouse": "仓库",
        "office tour": "办公室参观",
    },
    "de": {  # German
        "automatic door": "automatische Tür",
        "glass door": "Glastür",
        "entrance": "Eingang",
        "warehouse": "Lager",
        "walking tour": "Stadtrundgang",
    },
    "es": {  # Spanish
        "automatic door": "puerta automática",
        "glass door": "puerta de vidrio",
        "entrance": "entrada",
        "walking tour": "recorrido a pie",
        "shopping mall": "centro comercial",
    },
    "ar": {  # Arabic
        "automatic door": "باب أوتوماتيكي",
        "glass door": "باب زجاجي",
        "entrance": "مدخل",
    },
    "hi": {  # Hindi
        "automatic door": "ऑटोमैटिक दरवाज़ा",
        "glass door": "कांच का दरवाज़ा",
        "warehouse": "गोदाम",
    },
}

QUALITY_SUFFIXES = ["4k", "4k 60fps", "hd", "gopro"]
YEAR_SUFFIXES = ["2024", "2025", "2026"]


def expand_queries(queries: list[SearchQuery], languages: bool = True,
                   quality: bool = True, years: bool = False) -> list[SearchQuery]:
    """Generate variations of existing queries for broader coverage."""
    expanded = []

    for q in queries:
        # Only expand original queries, skip already-expanded ones
        if q.is_variation:
            continue

        # Language expansions (only for high-yield, high-priority queries)
        if languages and q.priority_rank <= 4 and q.yield_ratio >= 0.25:
            for lang_code, translations in LANGUAGE_EXPANSIONS.items():
                for eng_phrase, translated in translations.items():
                    if eng_phrase.lower() in q.term.lower():
                        new_term = q.term.lower().replace(eng_phrase.lower(), translated)
                        # Avoid duplicates
                        if new_term != q.term.lower():
                            expanded.append(SearchQuery(
                                term=new_term,
                                priority=q.priority,
                                priority_rank=q.priority_rank,
                                category=q.category,
                                sub_category=q.sub_category,
                                est_videos=max(10, q.est_videos // 3),
                                yield_ratio=q.yield_ratio,
                                est_images=max(3, q.est_images / 3),
                                notes=f"[{lang_code} variant] {q.notes}",
                                is_variation=True,
                                variation_type=f"lang_{lang_code}",
                            ))
                            break  # One translation per language per query

        # Quality suffix expansions (for walking tours, city content)
        if quality and q.priority_rank <= 3 and q.est_videos >= 50:
            # Only add "4k" if not already present
            term_lower = q.term.lower()
            if "4k" not in term_lower and "hd" not in term_lower:
                expanded.append(SearchQuery(
                    term=f"{q.term} 4k",
                    priority=q.priority,
                    priority_rank=q.priority_rank,
                    category=q.category,
                    sub_category=q.sub_category,
                    est_videos=max(10, q.est_videos // 2),
                    yield_ratio=min(1.0, q.yield_ratio * 1.1),
                    est_images=max(3, q.est_images / 2),
                    notes=f"[4K variant] {q.notes}",
                    is_variation=True,
                    variation_type="quality_4k",
                ))

        # Year suffixes (for recent content)
        if years and q.priority_rank <= 2:
            for year in YEAR_SUFFIXES:
                if year not in q.term:
                    expanded.append(SearchQuery(
                        term=f"{q.term} {year}",
                        priority=q.priority,
                        priority_rank=q.priority_rank,
                        category=q.category,
                        sub_category=q.sub_category,
                        est_videos=max(5, q.est_videos // 4),
                        yield_ratio=q.yield_ratio,
                        est_images=max(2, q.est_images / 4),
                        notes=f"[{year} variant] {q.notes}",
                        is_variation=True,
                        variation_type=f"year_{year}",
                    ))

    return expanded


# ─────────────────────────────────────────────
# Filters
# ─────────────────────────────────────────────
def filter_queries(queries: list[SearchQuery],
                   priorities: Optional[list[str]] = None,
                   min_yield: float = 0.0,
                   min_videos: int = 0,
                   sub_categories: Optional[list[str]] = None,
                   exclude_variations: bool = False) -> list[SearchQuery]:
    filtered = queries
    if priorities:
        pri_set = {p.upper() for p in priorities}
        filtered = [q for q in filtered if q.priority in pri_set]
    if min_yield > 0:
        filtered = [q for q in filtered if q.yield_ratio >= min_yield]
    if min_videos > 0:
        filtered = [q for q in filtered if q.est_videos >= min_videos]
    if sub_categories:
        sc_lower = {s.lower() for s in sub_categories}
        filtered = [q for q in filtered if q.sub_category.lower() in sc_lower]
    if exclude_variations:
        filtered = [q for q in filtered if not q.is_variation]
    return filtered


def sort_queries(queries: list[SearchQuery], by: str = "priority") -> list[SearchQuery]:
    if by == "priority":
        return sorted(queries, key=lambda q: (q.priority_rank, -q.est_images))
    elif by == "yield":
        return sorted(queries, key=lambda q: -q.yield_ratio)
    elif by == "images":
        return sorted(queries, key=lambda q: -q.est_images)
    elif by == "videos":
        return sorted(queries, key=lambda q: -q.est_videos)
    return queries


# ─────────────────────────────────────────────
# Export functions
# ─────────────────────────────────────────────
def export_json(queries: list[SearchQuery], filepath: str):
    data = {
        "metadata": {
            "total_queries": len(queries),
            "total_est_videos": sum(q.est_videos for q in queries),
            "total_est_images": sum(q.est_images for q in queries),
            "priorities": sorted(set(q.priority for q in queries)),
            "categories": sorted(set(q.category for q in queries)),
            "original_queries": sum(1 for q in queries if not q.is_variation),
            "variation_queries": sum(1 for q in queries if q.is_variation),
        },
        "queries": [asdict(q) for q in queries],
    }
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print(f"  → JSON saved: {filepath} ({len(queries)} queries)")


def export_csv(queries: list[SearchQuery], filepath: str):
    fieldnames = [
        "term", "priority", "priority_rank", "category", "sub_category",
        "est_videos", "yield_ratio", "est_images", "notes",
        "youtube_url", "youtube_api_query", "is_variation", "variation_type",
    ]
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for q in queries:
            writer.writerow(asdict(q))
    print(f"  → CSV saved: {filepath} ({len(queries)} queries)")


def export_urls(queries: list[SearchQuery], filepath: str):
    """Export just YouTube search URLs, one per line — easy to feed into scrapers."""
    with open(filepath, "w", encoding="utf-8") as f:
        for q in queries:
            f.write(f"{q.youtube_url}\n")
    print(f"  → URLs saved: {filepath} ({len(queries)} URLs)")


def export_api_queries(queries: list[SearchQuery], filepath: str):
    """Export as JSON array of YouTube Data API v3 search params."""
    api_params = []
    for q in queries:
        params = q.build_api_params()
        params["_meta"] = {
            "priority": q.priority,
            "category": q.category,
            "expected_yield": q.yield_ratio,
        }
        api_params.append(params)
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(api_params, f, indent=2, ensure_ascii=False)
    print(f"  → API params saved: {filepath} ({len(api_params)} queries)")


def export_batch_script(queries: list[SearchQuery], filepath: str):
    """Export a yt-dlp compatible batch search file."""
    with open(filepath, "w", encoding="utf-8") as f:
        f.write("# yt-dlp batch search queries for door detection data collection\n")
        f.write("# Usage: yt-dlp --batch-file this_file.txt --max-downloads 50\n")
        f.write("#        --write-info-json --skip-download (for metadata only)\n\n")
        for q in queries:
            f.write(f"ytsearch50:{q.term}\n")
    print(f"  → yt-dlp batch file saved: {filepath} ({len(queries)} searches)")


# ─────────────────────────────────────────────
# Statistics
# ─────────────────────────────────────────────
def print_stats(queries: list[SearchQuery]):
    print("\n" + "=" * 70)
    print("  DOOR DETECTION DATA COLLECTION — QUERY STATISTICS")
    print("=" * 70)

    originals = [q for q in queries if not q.is_variation]
    variations = [q for q in queries if q.is_variation]

    print(f"\n  Total queries:      {len(queries)}")
    print(f"  Original queries:   {len(originals)}")
    print(f"  Variation queries:  {len(variations)}")
    print(f"  Est. total videos:  {sum(q.est_videos for q in queries):,}")
    print(f"  Est. good images:   {sum(q.est_images for q in queries):,.0f}")

    print(f"\n  {'Priority':<8} {'Category':<35} {'Queries':>8} {'Videos':>10} {'Images':>10} {'Avg Yield':>10}")
    print("  " + "-" * 81)

    by_priority = {}
    for q in queries:
        key = (q.priority, q.priority_rank, q.category)
        if key not in by_priority:
            by_priority[key] = []
        by_priority[key].append(q)

    for (pri, rank, cat), qs in sorted(by_priority.items(), key=lambda x: x[0][1]):
        total_v = sum(q.est_videos for q in qs)
        total_i = sum(q.est_images for q in qs)
        avg_y = sum(q.yield_ratio for q in qs) / len(qs) if qs else 0
        print(f"  {pri:<8} {cat:<35} {len(qs):>8} {total_v:>10,} {total_i:>10,.0f} {avg_y:>9.1%}")

    print("  " + "-" * 81)
    total_v = sum(q.est_videos for q in queries)
    total_i = sum(q.est_images for q in queries)
    avg_y = sum(q.yield_ratio for q in queries) / len(queries) if queries else 0
    print(f"  {'TOTAL':<8} {'':<35} {len(queries):>8} {total_v:>10,} {total_i:>10,.0f} {avg_y:>9.1%}")

    # Variation breakdown
    if variations:
        print(f"\n  Variation breakdown:")
        var_types = {}
        for q in variations:
            var_types[q.variation_type] = var_types.get(q.variation_type, 0) + 1
        for vt, count in sorted(var_types.items()):
            print(f"    {vt:<20} {count:>5} queries")

    # Target analysis
    target = 100_000
    print(f"\n  Target:             {target:,} images")
    print(f"  Current estimate:   {total_i:,.0f} images")
    gap = target - total_i
    if gap > 0:
        print(f"  Gap:                {gap:,.0f} images")
        print(f"  → Need ~{int(gap / 0.3):,} more videos at avg 0.3 yield")
        print(f"  → Or ~{int(gap / 0.35 / 50):,} more search queries at 50 vids × 0.35 yield")
    else:
        print(f"  Surplus:            {abs(gap):,.0f} images (buffer for quality filtering)")
    print()


# ─────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="Generate YouTube search queries from door detection data collection plan"
    )
    parser.add_argument("--excel", default="door_data_collection_plan.xlsx",
                        help="Path to Excel file (default: door_data_collection_plan.xlsx)")
    parser.add_argument("--output-dir", default="./output",
                        help="Output directory (default: ./output)")
    parser.add_argument("--priority", nargs="+", default=None,
                        help="Filter by priority (e.g., P1 P3 P4)")
    parser.add_argument("--min-yield", type=float, default=0.0,
                        help="Minimum yield ratio filter (e.g., 0.4)")
    parser.add_argument("--min-videos", type=int, default=0,
                        help="Minimum estimated videos filter")
    parser.add_argument("--expand", action="store_true",
                        help="Add language/quality/year variations")
    parser.add_argument("--expand-languages", action="store_true",
                        help="Add language variations only")
    parser.add_argument("--expand-quality", action="store_true",
                        help="Add 4K quality variants only")
    parser.add_argument("--expand-years", action="store_true",
                        help="Add year suffix variants only")
    parser.add_argument("--max-results", type=int, default=50,
                        help="Max YouTube results per query for API params")
    parser.add_argument("--sort", choices=["priority", "yield", "images", "videos"],
                        default="priority", help="Sort order")
    parser.add_argument("--format", choices=["csv", "json", "both", "urls", "api", "ytdlp", "all"],
                        default="all", help="Output format")
    parser.add_argument("--stats", action="store_true",
                        help="Print statistics only (no file output)")
    parser.add_argument("--originals-only", action="store_true",
                        help="Exclude variation queries from output")

    args = parser.parse_args()

    # Read base queries
    print(f"\n📖 Reading queries from: {args.excel}")
    queries = read_queries_from_excel(args.excel)
    print(f"   Loaded {len(queries)} base queries")

    # Expand if requested
    if args.expand or args.expand_languages or args.expand_quality or args.expand_years:
        langs = args.expand or args.expand_languages
        qual = args.expand or args.expand_quality
        yrs = args.expand or args.expand_years
        expanded = expand_queries(queries, languages=langs, quality=qual, years=yrs)
        queries.extend(expanded)
        print(f"   + {len(expanded)} variations → {len(queries)} total")

    # Filter
    queries = filter_queries(
        queries,
        priorities=args.priority,
        min_yield=args.min_yield,
        min_videos=args.min_videos,
        exclude_variations=args.originals_only,
    )
    print(f"   After filters: {len(queries)} queries")

    # Sort
    queries = sort_queries(queries, by=args.sort)

    # Stats
    if args.stats:
        print_stats(queries)
        return

    # Always print stats
    print_stats(queries)

    # Export
    os.makedirs(args.output_dir, exist_ok=True)
    fmt = args.format

    if fmt in ("json", "both", "all"):
        export_json(queries, os.path.join(args.output_dir, "youtube_queries.json"))

    if fmt in ("csv", "both", "all"):
        export_csv(queries, os.path.join(args.output_dir, "youtube_queries.csv"))

    if fmt in ("urls", "all"):
        export_urls(queries, os.path.join(args.output_dir, "youtube_search_urls.txt"))

    if fmt in ("api", "all"):
        export_api_queries(queries, os.path.join(args.output_dir, "youtube_api_params.json"))

    if fmt in ("ytdlp", "all"):
        export_batch_script(queries, os.path.join(args.output_dir, "ytdlp_batch.txt"))

    print(f"\n✅ Done! Files in: {os.path.abspath(args.output_dir)}/")


if __name__ == "__main__":
    main()
